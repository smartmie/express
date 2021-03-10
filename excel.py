import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Border,Side,Font,Alignment
import os
import get_express
import re
import datetime
def get_path_list(path):
    list_file = os.listdir(path)
    return list_file

def main(filename,index=0):
    excel = openpyxl.load_workbook(filename)
    sheet_name = excel.sheetnames
    sheet = excel.get_sheet_by_name(sheet_name[index])
    return sheet,excel

def get_time():
    date_time = datetime.datetime()
    time_yeas = date_time.year
    time_month = date_time.month
    time_day = date_time.day
    result = "{0}-{1}-{2}".format(time_yeas,time_month,time_day)
    return result

def set_file(name):
    now_path = os.getcwd()
    os.makedirs(now_path+"\\date\\"+name)
    os.makedirs(now_path+"\\date\\"+name+"\\image")

def red_column(sheet,index = 'L'):
    col = sheet[index]
    express = [e.value for e in col]
    return express
    
def red_row(sheet,index = '1'):
    col = sheet[index]
    express = [e.value for e in col]
    return express

def wri_excel(sheet, index = 1, y = 'M',text = "", image = None ):
    xy = '{0}{1}'.format(y,index)
    sheet[xy] = text
    sheet[xy].font = Font(name='等线',size= '16')
    sheet.row_dimensions[index].height = 100
    border = Border(left=Side(border_style='thin',color='ffffff'),right=Side(border_style='thin',color='ffffff'),top=Side(border_style='thin',color='ffffff'),bottom=Side(border_style='thin',color='ffffff'))
    sheet[xy].border = border
    sheet[xy].alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
    
    if image != None :
        image.height = 100
        image.width = 150
        sheet.add_image(image,"{0}{1}".format(y,index))

def set_color(sheet,x):
    style = openpyxl.styles.PatternFill("solid", fgColor="00FF00")
    for y in range(1,sheet.max_column):
        sheet.cell(x,y).fill = style
        pass
    pass

def list_image(list_path):
    list_red_image = {}
    for im in get_path_list(list_path):
        print(im)
        key = re.findall(r'\d+',im)[0]
        list_red_image.update({str(key):Image(list_path+"\\"+im)})
    return list_red_image
def red_image(path):
    return Image(path)

if __name__ == "__main__":
    pass
    # sheet , excel = main("one.xlsx",0)
    # set_color(sheet,2)
    # now_path = os.getcwd()+"\\date\\"
    # list_path = get_path_list(now_path)
    

    # # print(red_express(sheet))
    # # print(red_express(sheet,'B'))
    # # data_json = get_express.get_data('YT5106522924480')
    # # print(data_json['data'][0]['context'])
    # # wri_express(sheet,3,data_json['data'][0]['context'])
    # excel.save("b.xlsx")
    # pass
    